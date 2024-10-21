import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def generate_excel_yearly_report():
    current_year = datetime.now().year
    file_name = f"Jahreswareneingang_{current_year}.xlsx"
    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'results', file_name))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jahreswareneingang"
    sheet.append(["Warennummer", "Ursprungsland", "Zielland", "Gewicht", "Preis", "Abschnitt", "Abschnittsbeschreibung", "Infrastat-Beschreibung"])

    for month in range(1, 13):
        month_str = f"{month:02d}"  
        monthly_file_name = f"Wareneingang_{month_str}{current_year}.xlsx"
        monthly_file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'results', monthly_file_name))

        if os.path.exists(monthly_file_path):
            print(f"Monatsbericht für {month_str}/{current_year} gefunden: {monthly_file_name}")
            monthly_workbook = load_workbook(monthly_file_path)
            monthly_sheet = monthly_workbook.active

            for row in monthly_sheet.iter_rows(min_row=2, values_only=True):
                INFRASTAT_NUMBER, ORIGIN, DESTINATION, WEIGHT, PRICE, TAX_CHAPTER, TAX_CHAPTER_DESCRIPTION, INFRASTAT_DESCRIPTION = row


                origin_for_comparison = ORIGIN.strip().lower() if ORIGIN else ""
                destination_for_comparison = DESTINATION.strip().lower() if DESTINATION else ""

                found_match = False
                for yearly_row_idx, yearly_row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                    yearly_warennummer = int(yearly_row[0].value)
                    yearly_ursprungsland = yearly_row[1].value.strip().lower() if yearly_row[1].value else ""
                    yearly_zielland = yearly_row[2].value.strip().lower() if yearly_row[2].value else ""

                    if (yearly_warennummer == int(INFRASTAT_NUMBER) and 
                        yearly_ursprungsland == origin_for_comparison and 
                        yearly_zielland == destination_for_comparison):

                        found_match = True
                        new_weight = yearly_row[3].value + WEIGHT
                        new_price = (yearly_row[4].value + PRICE) / 2 if yearly_row[4].value != 0 and PRICE != 0 else 0

                        sheet.cell(row=yearly_row_idx, column=4, value=new_weight)  
                        sheet.cell(row=yearly_row_idx, column=5, value=new_price)   
                        break

                if not found_match:
                    sheet.append([INFRASTAT_NUMBER, ORIGIN, DESTINATION, WEIGHT, PRICE, TAX_CHAPTER, TAX_CHAPTER_DESCRIPTION, INFRASTAT_DESCRIPTION])


    workbook.save(file_path)
    print(f"Der Jahresbericht '{file_name}' wurde erfolgreich erstellt und alle Monatsberichte zusammengeführt.")

if __name__ == "__main__":
    generate_excel_yearly_report()