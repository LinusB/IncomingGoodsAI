import os
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'classification/config')))
from config import INFRASTAT_NUMBER, ORIGIN, DESTINATION, WEIGHT, PRICE, TAX_CHAPTER, TAX_CHAPTER_DESCRIPTION, INFRASTAT_DESCRIPTION
from openpyxl import Workbook, load_workbook
from datetime import datetime

def generate_excel_report():
    current_time = datetime.now()
    file_name = f"Wareneingang_{current_time.strftime('%m%Y')}.xlsx"
    file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'results', file_name))

    # Checking if file exists
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Wareneingang"
        sheet.append(["Warennummer", "Ursprungsland", "Zielland", "Gewicht", "Preis", "Abschnitt", "Abschnittsbeschreibung", "Infrastat-Beschreibung"])

    new_row = [INFRASTAT_NUMBER, ORIGIN, DESTINATION, WEIGHT, PRICE, TAX_CHAPTER, TAX_CHAPTER_DESCRIPTION, INFRASTAT_DESCRIPTION]

    found_match = False

    # Checking for matches...
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        existing_warennummer = int(row[0].value) if row[0].value is not None else None
        existing_ursprungsland = row[1].value.strip().lower() if row[1].value else ""  
        existing_zielland = row[2].value.strip().lower() if row[2].value else ""  

        if (existing_warennummer == int(INFRASTAT_NUMBER) and 
            existing_ursprungsland == ORIGIN.strip().lower() and 
            existing_zielland == DESTINATION.strip().lower()):
            
            # Match found, update existing row
            found_match = True
            new_weight = row[3].value + WEIGHT

            if row[4].value != 0 and PRICE != 0:
                new_price = (row[4].value + PRICE) / 2
            else:
                new_price = 0

            # Update the cells in the found row
            sheet.cell(row=row_idx, column=4, value=new_weight)
            sheet.cell(row=row_idx, column=5, value=new_price)

            break  

    # No match found, just appending a new row
    if not found_match:
        sheet.append([INFRASTAT_NUMBER, ORIGIN, DESTINATION, WEIGHT, PRICE, TAX_CHAPTER, TAX_CHAPTER_DESCRIPTION, INFRASTAT_DESCRIPTION])

    # Saving the file
    workbook.save(file_path)
    print(f"Die Datei '{file_name}' wurde erfolgreich erstellt oder aktualisiert.")


if __name__ == "__main__":
    generate_excel_report()
