import os
import subprocess
import pandas as pd
from flask import Flask, render_template, send_from_directory, request, make_response
from flask_socketio import SocketIO
from openpyxl import load_workbook
from datetime import datetime
import calendar

app = Flask(__name__)
socketio = SocketIO(app)

script_dir = os.path.dirname(os.path.abspath(__file__))
results_dir = os.path.join(script_dir, '..', 'results')
image_capturing_path = os.path.join(script_dir, '..', 'capturing', 'imageCapturing.py')
image_classification_path = os.path.join(script_dir, '..', 'classification', 'imageClassification.py')
generate_excel_monthly_report_path = os.path.join(script_dir, '..', 'creation', 'generate_excel_monthly_report.py')
generate_excel_yearly_report_path = os.path.join(script_dir, '..', 'creation', 'generate_excel_yearly_report.py')

current_process_dir = os.path.join(script_dir, './currentProcess')  # Ordnerpfad für currentProcess

@app.route('/currentProcess/<path:filename>')
def serve_current_process_file(filename):
    return send_from_directory(current_process_dir, filename)

def run_script(script_path):
    process = subprocess.Popen(['python', script_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    for stdout_line in iter(process.stdout.readline, ""):
        socketio.emit('status_update', {'message': stdout_line})
    process.stdout.close()
    process.wait()
    for stderr_line in iter(process.stderr.readline, ""):
        socketio.emit('status_update', {'message': stderr_line})
    process.stderr.close()
    return process.returncode

def get_available_months():
    files = os.listdir(results_dir)
    months = sorted(set(f.split('_')[1].split('.')[0] for f in files if f.startswith('Wareneingang_')))
    return months

@app.route('/start_capture', methods=['GET'])
def start_capture():
    try:
        # Führe ImageCapturing.py aus
        returncode_capture = run_script(image_capturing_path)
        
        if returncode_capture == 0:
            # Führe imageClassification.py aus
            returncode_classification = run_script(image_classification_path)
            
            if returncode_classification == 0:
                socketio.emit('status_update', {'message': 'Bildaufnahme und Klassifizierung erfolgreich'})
                returncode_report = run_script(generate_excel_monthly_report_path)
                if returncode_report == 0:
                    socketio.emit('status_update', {'message': 'Monatlicher Excel-Bericht erfolgreich generiert'})
                else:
                    socketio.emit('status_update', {'message': 'Fehler beim Generieren des monatlichen Excel-Berichts'})
            else:
                socketio.emit('status_update', {'message': 'Fehler bei der Klassifizierung'})
        else:
            socketio.emit('status_update', {'message': 'Fehler bei der Bildaufnahme'})
    except Exception as e:
        socketio.emit('status_update', {'message': str(e)})

# Funktion, um nur die gewünschten Spalten der letzten 5 Einträge zu extrahieren
def get_last_five_entries(file_path):
    if not os.path.exists(file_path):
        return []

    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Wir interessieren uns nur für die Spalten:
    # - Warennummer (Spalte 1)
    # - Abschnittsbeschreibung (Spalte 7)
    # - Infrastat-Beschreibung (Spalte 8)
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    recent_entries = rows[-5:]  # Die letzten 5 Zeilen auswählen
    
    # Nur die gewünschten Spalten extrahieren
    filtered_entries = [(row[0], row[6], row[7]) for row in recent_entries]
    
    return filtered_entries

# Route für die Hauptseite
@app.route('/')
def index():
    response = make_response(render_template('index.html'))
    
    # Setze die Header, um das Caching zu deaktivieren
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'

    return response

# Route für die Report-Seite
@app.route('/report')
def report():
    excel_file_path = os.path.join(script_dir, '..', 'results', f"Wareneingang_{datetime.now().strftime('%m%Y')}.xlsx")
    recent_entries = get_last_five_entries(excel_file_path)
    
    # Rufe die verfügbaren Monate für das Dropdown-Menü ab
    available_months = get_available_months()

    response = make_response(render_template('report.html', entries=recent_entries, months=available_months))
    
    # Setze die Header, um das Caching zu deaktivieren
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

# Filter zur Umwandlung von numerischen Monatswerten in Monatsnamen
@app.template_filter('format_month')
def format_month(value):
    try:
        return calendar.month_name[value]
    except:
        return "Unbekannter Monat"

# Route für den Jahresreport (Excel-Datei des aktuellen Jahres)
@app.route('/download_yearly_report')
def download_yearly_report():
    returncode_report = run_script(generate_excel_yearly_report_path)
    if returncode_report == 0:
        socketio.emit('status_update', {'message': 'Monatlicher Excel-Bericht erfolgreich generiert'})
    else:
        socketio.emit('status_update', {'message': 'Fehler beim Generieren des monatlichen Excel-Berichts'})
    # Generiere den Dateinamen basierend auf dem aktuellen Jahr
    year_file_name = f"Jahreswareneingang_{datetime.now().strftime('%Y')}.xlsx"
    file_path = os.path.join(script_dir, '..', 'results')  # Verzeichnis, in dem sich die Datei befindet
    return send_from_directory(directory=file_path, path=year_file_name, as_attachment=True)

# Route für das Herunterladen eines ausgewählten Monatsberichts
@app.route('/download_report/<month_year>', methods=['GET'])
def download_report(month_year):
    month_file_name = f"Wareneingang_{month_year}.xlsx"
    file_path = os.path.join(script_dir, '..', 'results')  # Verzeichnis, in dem sich die Datei befindet
    print(f"Trying to download: {month_file_name}")
    
    # Überprüfe, ob die Datei existiert
    if not os.path.exists(os.path.join(file_path, month_file_name)):
        return "Report für den ausgewählten Monat nicht gefunden.", 404

    # Datei herunterladen
    return send_from_directory(directory=file_path, path=month_file_name, as_attachment=True)

if __name__ == '__main__':
    socketio.run(app, debug=True)

